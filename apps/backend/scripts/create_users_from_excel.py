#!/usr/bin/env python3
"""
엑셀에서 사용자 추출 및 DB 생성 스크립트

엑셀 파일의 "현 사용자" 컬럼에서 실제 사용자 이름을 추출하여 DB에 생성합니다.
"""

import argparse
import asyncio
import re
import sys
import uuid
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.database import AsyncSessionLocal
from src.models.user import User, UserRole
from src.utils.security import hash_password


class UserCreationStats:
    """사용자 생성 통계"""

    def __init__(self):
        self.total_found = 0
        self.created_count = 0
        self.skipped_count = 0
        self.existing_count = 0
        self.skipped_names = []
        self.created_names = []

    def print_summary(self):
        print("\n" + "=" * 60)
        print("📊 사용자 생성 결과")
        print("=" * 60)
        print(f"발견된 이름: {self.total_found}개")
        print(f"✅ 생성됨: {self.created_count}개")
        print(f"⏭️  이미 존재: {self.existing_count}개")
        print(f"❌ 제외됨: {self.skipped_count}개")

        if self.created_names:
            print(f"\n✅ 생성된 사용자 샘플 (처음 20명):")
            for name in self.created_names[:20]:
                print(f"   - {name}")
            if len(self.created_names) > 20:
                print(f"   ... 외 {len(self.created_names) - 20}명")

        if self.skipped_names:
            print(f"\n❌ 제외된 이름 샘플 (처음 20개):")
            for name in self.skipped_names[:20]:
                print(f"   - {name}")
            if len(self.skipped_names) > 20:
                print(f"   ... 외 {len(self.skipped_names) - 20}개")


def clean_string(value: Any) -> str | None:
    """문자열 정제"""
    if value is None:
        return None
    value_str = str(value).strip()
    return value_str if value_str else None


def is_valid_user_name(name: str) -> bool:
    """
    실제 사용자 이름인지 검증

    제외 대상:
    - 빈 값, None, '-', 'nan'
    - 상태 값: "폐기", "폐기완료", "불용예정", "창고", "보관" 등
    - 장소/부서: "회의실", "서버실", "개발실", "팀", "실" 등이 포함된 경우
    - 기타: "서버", "공용", "대여", "전시회", "TBD" 등
    """
    if not name:
        return False

    # 제외할 키워드 리스트
    exclude_keywords = [
        '회의실', '서버실', '개발실', '서버', '공용', '대여',
        '폐기', '불용', '창고', '보관', '재고',
        '전시회', 'TBD', 'Cloud', 'PMS', 'AX', 'SDx', 'SQA',
        '확인필', '미확인', '지급장비', '대여용', '일반장비',
        '품질기술팀', '사업개발'
    ]

    # 제외할 정확한 값들
    exclude_exact = {
        '-', 'nan', 'NaN', 'N/A', 'n/a', '없음', '미정', '노후'
    }

    # 정확히 일치하는 경우 제외
    if name in exclude_exact:
        return False

    # 키워드 포함 여부 확인
    for keyword in exclude_keywords:
        if keyword in name:
            return False

    # 숫자로 시작하는 경우 제외 (예: "10-6회의실", "4층 서버실")
    if re.match(r'^\d', name):
        return False

    # 괄호 안에 숫자만 있는 경우 제외 (예: "(154서버)")
    if re.match(r'^.*\(\d+.*\).*$', name):
        return False

    # 영문 대문자로만 구성된 경우 제외 (팀명 등)
    if re.match(r'^[A-Z]+$', name):
        return False

    # 한글 이름 패턴 확인 (2-4자)
    korean_name_pattern = r'^[가-힣]{2,4}$'
    if re.match(korean_name_pattern, name):
        return True

    # 영문 이름 뒤에 알파벳 하나 (강동훈B 등) - 유효한 이름으로 처리
    if re.match(r'^[가-힣]{2,4}[A-Z]$', name):
        return True

    # 그 외의 경우는 제외
    return False


def extract_users_from_excel(excel_path: str) -> set[str]:
    """엑셀 파일에서 사용자 이름 추출"""
    print("📂 Excel 파일 로드 중...")
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    print(f"✅ {len(wb.sheetnames)}개 시트 발견")

    user_names = set()
    sheet_names = ['데스크탑(11)', '노트북(12)', '모니터(14)']

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  시트를 찾을 수 없음: {sheet_name}")
            continue

        print(f"\n📄 {sheet_name} 시트 분석 중...")
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        sheet_users = set()
        for row in rows:
            current_user = row[1] if len(row) > 1 else None
            if current_user:
                user_str = clean_string(current_user)
                if user_str and is_valid_user_name(user_str):
                    user_names.add(user_str)
                    sheet_users.add(user_str)

        print(f"  ✅ 유효한 사용자: {len(sheet_users)}명")

    print(f"\n✅ 총 고유 사용자: {len(user_names)}명")
    return user_names


async def create_user(
    db: AsyncSession,
    name: str,
    email: str,
    password: str = "user123!",
    role: UserRole = UserRole.EMPLOYEE,
) -> User:
    """사용자 생성"""
    user = User(
        id=str(uuid.uuid4()),
        name=name,
        email=email,
        password_hash=hash_password(password),
        role=role,
        is_active=True,
        is_verified=False,
    )
    db.add(user)
    await db.flush()
    return user


async def get_user_by_email(db: AsyncSession, email: str) -> User | None:
    """이메일로 사용자 조회"""
    result = await db.execute(select(User).where(User.email == email))
    return result.scalar_one_or_none()


async def get_user_by_name(db: AsyncSession, name: str) -> User | None:
    """이름으로 사용자 조회"""
    result = await db.execute(select(User).where(User.name == name))
    return result.scalar_one_or_none()


def generate_email(name: str, counter: int = 0) -> str:
    """
    이름에서 이메일 생성

    예:
    - "박재현" -> "박재현@suresoft.com"
    - "박재현" (중복) -> "박재현2@suresoft.com"
    - "강동훈B" -> "강동훈B@suresoft.com"
    """
    if counter == 0:
        return f"{name}@suresoft.com"
    else:
        return f"{name}{counter}@suresoft.com"


async def create_users_from_names(
    db: AsyncSession,
    user_names: set[str],
    stats: UserCreationStats,
    dry_run: bool = False,
) -> None:
    """사용자 이름 목록에서 DB에 사용자 생성"""
    print("\n" + "=" * 60)
    print("👥 사용자 생성 중...")
    print("=" * 60)

    stats.total_found = len(user_names)

    for name in sorted(user_names):
        try:
            # 이름으로 이미 존재하는지 확인
            existing = await get_user_by_name(db, name)
            if existing:
                stats.existing_count += 1
                continue

            # 이메일 생성 (중복 처리)
            counter = 0
            email = generate_email(name, counter)

            while await get_user_by_email(db, email):
                counter += 1
                email = generate_email(name, counter)
                if counter > 100:  # 안전 장치
                    print(f"   ⚠️  {name}: 이메일 생성 실패 (너무 많은 중복)")
                    stats.skipped_count += 1
                    stats.skipped_names.append(name)
                    break
            else:
                # 사용자 생성
                if not dry_run:
                    await create_user(db, name, email)

                stats.created_count += 1
                stats.created_names.append(name)

                # 진행 상황 출력 (50개마다)
                if stats.created_count % 50 == 0:
                    print(f"   ✅ {stats.created_count}명 생성 완료...")

        except Exception as e:
            print(f"   ❌ {name}: {str(e)}")
            stats.skipped_count += 1
            stats.skipped_names.append(name)


async def create_users(
    excel_path: str,
    dry_run: bool = False,
    verify_only: bool = False,
) -> None:
    """사용자 생성 메인 함수"""
    print("=" * 60)
    print("👥 엑셀에서 사용자 생성")
    print("=" * 60)
    print(f"파일: {excel_path}")
    print(f"모드: {'검증만' if verify_only else '드라이런' if dry_run else '실제 생성'}")
    print("=" * 60)

    # 파일 존재 확인
    if not Path(excel_path).exists():
        print(f"❌ 파일을 찾을 수 없습니다: {excel_path}")
        return

    # 엑셀에서 사용자 추출
    user_names = extract_users_from_excel(excel_path)

    if verify_only:
        print("\n📋 추출된 사용자 목록:")
        for i, name in enumerate(sorted(user_names), 1):
            print(f"  {i}. {name}")
        print(f"\n✅ 총 {len(user_names)}명")
        return

    # DB에 사용자 생성
    stats = UserCreationStats()

    async with AsyncSessionLocal() as db:
        try:
            await create_users_from_names(db, user_names, stats, dry_run)

            # 커밋 또는 롤백
            if not dry_run:
                await db.commit()
                print("\n✅ 모든 변경사항 저장 완료")
            else:
                await db.rollback()
                print("\n🔍 드라이런 완료 (변경사항 저장 안함)")

            # 통계 출력
            stats.print_summary()

        except Exception as e:
            await db.rollback()
            print(f"\n❌ 사용자 생성 실패: {e}")
            raise


def main():
    parser = argparse.ArgumentParser(description="엑셀에서 사용자 추출 및 생성")
    parser.add_argument(
        "--file",
        default="자산관리 데이터(슈커톤).xlsx",
        help="Excel 파일 경로 (기본: 자산관리 데이터(슈커톤).xlsx)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="드라이런 모드 (실제 저장하지 않고 검증만)",
    )
    parser.add_argument(
        "--verify-only",
        action="store_true",
        help="추출된 사용자 목록만 확인 (DB 작업 없음)",
    )

    args = parser.parse_args()

    # 프로젝트 루트로 이동 (apps/backend -> suresoft-ams)
    project_root = Path(__file__).parent.parent.parent.parent
    excel_path = project_root / args.file

    asyncio.run(create_users(str(excel_path), args.dry_run, args.verify_only))


if __name__ == "__main__":
    main()
