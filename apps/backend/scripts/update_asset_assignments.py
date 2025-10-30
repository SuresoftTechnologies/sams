#!/usr/bin/env python3
"""
자산의 assigned_to 업데이트 스크립트

엑셀 데이터를 기반으로 기존 자산의 assigned_to를 업데이트합니다.
"""

import argparse
import asyncio
import sys
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.database import AsyncSessionLocal
from src.models.asset import Asset
from src.models.user import User


def clean_string(value: Any) -> str | None:
    """문자열 정제"""
    if value is None:
        return None
    value_str = str(value).strip()
    return value_str if value_str else None


async def get_user_by_name(db: AsyncSession, name: str) -> User | None:
    """사용자 이름으로 조회"""
    if not name:
        return None

    result = await db.execute(select(User).where(User.name == name))
    return result.scalar_one_or_none()


async def update_asset_assignment(
    db: AsyncSession,
    asset_tag: str,
    user_name: str | None,
) -> tuple[bool, str]:
    """자산의 assigned_to 업데이트"""
    # 자산 조회
    result = await db.execute(select(Asset).where(Asset.asset_tag == asset_tag))
    asset = result.scalar_one_or_none()

    if not asset:
        return False, "자산을 찾을 수 없음"

    # 사용자 조회
    if user_name:
        user = await get_user_by_name(db, user_name)
        if user:
            asset.assigned_to = user.id
            return True, f"할당: {user_name}"
        else:
            asset.assigned_to = None
            return False, f"사용자 없음: {user_name}"
    else:
        asset.assigned_to = None
        return True, "미할당"


async def update_assignments_from_excel(
    excel_path: str,
    dry_run: bool = False,
) -> None:
    """엑셀 데이터에서 자산 할당 업데이트"""
    print("=" * 60)
    print("🔄 자산 할당 업데이트")
    print("=" * 60)
    print(f"파일: {excel_path}")
    print(f"모드: {'드라이런 (실제 저장 안함)' if dry_run else '실제 업데이트'}")
    print("=" * 60)

    # 파일 존재 확인
    if not Path(excel_path).exists():
        print(f"❌ 파일을 찾을 수 없습니다: {excel_path}")
        return

    # Excel 파일 로드
    print("\n📂 Excel 파일 로드 중...")
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    print(f"✅ {len(wb.sheetnames)}개 시트 발견")

    # 통계
    total_rows = 0
    updated_count = 0
    not_found_count = 0
    no_user_count = 0
    unchanged_count = 0

    async with AsyncSessionLocal() as db:
        try:
            # 시트 처리
            sheet_names = ["데스크탑(11)", "노트북(12)", "모니터(14)"]

            for sheet_name in sheet_names:
                if sheet_name not in wb.sheetnames:
                    continue

                print(f"\n📄 {sheet_name} 시트 처리 중...")
                sheet = wb[sheet_name]
                rows = list(sheet.iter_rows(min_row=2, values_only=True))

                for row in rows:
                    total_rows += 1

                    # 데이터 추출
                    asset_tag = clean_string(row[0])  # 자산번호
                    current_user_name = clean_string(row[1])  # 현 사용자

                    if not asset_tag:
                        continue

                    # 할당 업데이트
                    success, message = await update_asset_assignment(
                        db, asset_tag, current_user_name
                    )

                    if success:
                        if "할당" in message:
                            updated_count += 1
                        else:
                            unchanged_count += 1
                    else:
                        if "찾을 수 없음" in message:
                            not_found_count += 1
                        elif "사용자 없음" in message:
                            no_user_count += 1

                    # 진행 상황 출력
                    if total_rows % 100 == 0:
                        print(f"   ✅ {total_rows}개 처리 중...")

                print(f"   완료: {len(rows)}개 처리")

            # 커밋 또는 롤백
            if not dry_run:
                await db.commit()
                print("\n✅ 모든 변경사항 저장 완료")
            else:
                await db.rollback()
                print("\n🔍 드라이런 완료 (변경사항 저장 안함)")

            # 통계 출력
            print("\n" + "=" * 60)
            print("📊 업데이트 결과")
            print("=" * 60)
            print(f"총 처리: {total_rows}개")
            print(f"✅ 할당 업데이트: {updated_count}개")
            print(f"⏭️  미할당: {unchanged_count}개")
            print(f"❌ 자산 없음: {not_found_count}개")
            print(f"❌ 사용자 없음: {no_user_count}개")

        except Exception as e:
            await db.rollback()
            print(f"\n❌ 업데이트 실패: {e}")
            raise


def main():
    parser = argparse.ArgumentParser(description="자산 할당 업데이트")
    parser.add_argument(
        "--file",
        default="자산관리 데이터(슈커톤).xlsx",
        help="Excel 파일 경로 (기본: 자산관리 데이터(슈커톤).xlsx)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="드라이런 모드 (실제 저장하지 않고 검증만)",
    )

    args = parser.parse_args()

    # 프로젝트 루트로 이동
    project_root = Path(__file__).parent.parent.parent.parent
    excel_path = project_root / args.file

    asyncio.run(update_assignments_from_excel(str(excel_path), args.dry_run))


if __name__ == "__main__":
    main()
