#!/usr/bin/env python3
"""
Excel 데이터 마이그레이션 스크립트

실제 현장 데이터를 안전하게 데이터베이스로 마이그레이션합니다.
"""

import argparse
import asyncio
import sys
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.config import settings
from src.database import AsyncSessionLocal
from src.models.asset import Asset, AssetStatus, AssetGrade
from src.models.category import Category
from src.models.location import Location
from src.models.user import User
from src.utils.security import hash_password


class MigrationStats:
    """마이그레이션 통계"""

    def __init__(self):
        self.total_rows = 0
        self.success_count = 0
        self.skip_count = 0
        self.error_count = 0
        self.errors = []

    def add_success(self):
        self.success_count += 1

    def add_skip(self, reason: str):
        self.skip_count += 1
        print(f"   ⏭️  스킵: {reason}")

    def add_error(self, row_num: int, error: str):
        self.error_count += 1
        self.errors.append((row_num, error))
        print(f"   ❌ 오류 (행 {row_num}): {error}")

    def print_summary(self):
        print("\n" + "=" * 60)
        print("📊 마이그레이션 결과")
        print("=" * 60)
        print(f"총 처리: {self.total_rows}개")
        print(f"✅ 성공: {self.success_count}개")
        print(f"⏭️  스킵: {self.skip_count}개")
        print(f"❌ 실패: {self.error_count}개")

        if self.errors:
            print("\n❌ 오류 상세:")
            for row_num, error in self.errors[:10]:  # 최대 10개만 표시
                print(f"   행 {row_num}: {error}")
            if len(self.errors) > 10:
                print(f"   ... 그 외 {len(self.errors) - 10}개")


def clean_string(value: Any) -> str | None:
    """문자열 정제"""
    if value is None:
        return None
    value_str = str(value).strip()
    return value_str if value_str else None


def truncate_string(value: str | None, max_length: int) -> str | None:
    """문자열 길이 제한"""
    if not value:
        return value
    if len(value) > max_length:
        return value[: max_length - 3] + "..."
    return value


def parse_date(value: Any) -> datetime | None:
    """날짜 파싱"""
    if not value:
        return None

    if isinstance(value, datetime):
        return value

    try:
        # 다양한 날짜 형식 지원
        date_str = str(value).strip()
        for fmt in ["%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"]:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        return None
    except Exception:
        return None


def parse_price(value: Any) -> int | None:
    """가격 파싱"""
    if not value:
        return None
    try:
        # 콤마, 원화 기호 제거
        price_str = str(value).replace(",", "").replace("원", "").replace("₩", "").strip()
        return int(float(price_str))
    except Exception:
        return None


def determine_grade(purchase_date: datetime | None) -> AssetGrade:
    """구매 연도 기반 등급 자동 계산"""
    if not purchase_date:
        return AssetGrade.C

    purchase_year = purchase_date.year
    current_year = datetime.now().year

    # A급: 2022~2025년
    if purchase_year >= 2022:
        return AssetGrade.A
    # B급: 2018~2021년
    elif purchase_year >= 2018:
        return AssetGrade.B
    # C급: ~2017년
    else:
        return AssetGrade.C


def map_status(status_str: str | None) -> AssetStatus:
    """상태 매핑 - 실제 엑셀 데이터 기준"""
    if not status_str:
        return AssetStatus.STOCK

    # 대괄호 제거
    status_clean = status_str.replace("[", "").replace("]", "").strip()

    status_map = {
        "지급장비": AssetStatus.ISSUED,
        "대여용": AssetStatus.LOANED,
        "일반장비": AssetStatus.GENERAL,
        "재고": AssetStatus.STOCK,
        "서버실": AssetStatus.SERVER_ROOM,
        "불용": AssetStatus.DISPOSED,
    }

    return status_map.get(status_clean, AssetStatus.STOCK)

    status_map = {
        "사용중": AssetStatus.ASSIGNED,
        "대여": AssetStatus.ASSIGNED,
        "대여중": AssetStatus.ASSIGNED,
        "보관": AssetStatus.AVAILABLE,
        "사용가능": AssetStatus.AVAILABLE,
        "유지보수": AssetStatus.MAINTENANCE,
        "수리중": AssetStatus.MAINTENANCE,
        "불용": AssetStatus.DISPOSED,
        "폐기": AssetStatus.DISPOSED,
    }

    for key, value in status_map.items():
        if key in status_str:
            return value

    return AssetStatus.AVAILABLE


async def get_or_create_category(
    db: AsyncSession, name: str, code: str
) -> Category | None:
    """카테고리 조회 또는 생성"""
    result = await db.execute(select(Category).where(Category.code == code))
    category = result.scalar_one_or_none()

    if not category:
        category = Category(
            id=str(uuid.uuid4()),
            name=name,
            code=code,
            description=f"{name} 카테고리",
            is_active=True,
        )
        db.add(category)
        await db.flush()

    return category


async def get_or_create_location(
    db: AsyncSession, site: str, building: str | None = None, floor: str | None = None
) -> Location | None:
    """위치 조회 또는 생성"""
    # 코드 생성
    code_parts = [site[:2].upper()]
    if building:
        code_parts.append(building[:4].upper())
    if floor:
        code_parts.append(floor.upper())
    code = "-".join(code_parts)

    result = await db.execute(select(Location).where(Location.code == code))
    location = result.scalar_one_or_none()

    if not location:
        # 이름 생성
        name_parts = [site]
        if building:
            name_parts.append(building)
        if floor:
            name_parts.append(floor)
        name = " ".join(name_parts)

        location = Location(
            id=str(uuid.uuid4()),
            name=name,
            code=code,
            site=site,
            building=building,
            floor=floor,
            is_active=True,
        )
        db.add(location)
        await db.flush()

    return location


async def get_user_by_name(db: AsyncSession, name: str) -> User | None:
    """사용자 이름으로 조회"""
    if not name:
        return None

    result = await db.execute(select(User).where(User.name == name))
    return result.scalar_one_or_none()


async def migrate_sheet(
    db: AsyncSession,
    sheet: Any,
    category_name: str,
    category_code: str,
    stats: MigrationStats,
    dry_run: bool = False,
) -> None:
    """시트 데이터 마이그레이션"""
    print(f"\n📄 {category_name} 시트 처리 중...")

    # 카테고리 생성
    category = await get_or_create_category(db, category_name, category_code)

    # 헤더 행 스킵 (2번째 행부터)
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    stats.total_rows += len(rows)

    for idx, row in enumerate(rows, start=2):
        try:
            # 필수 필드 확인
            asset_tag = clean_string(row[0])  # 자산번호 (asset_tag)
            if not asset_tag:
                stats.add_skip(f"행 {idx}: 자산번호 없음")
                continue

            # 중복 확인
            existing = await db.execute(
                select(Asset).where(Asset.asset_tag == asset_tag)
            )
            if existing.scalar_one_or_none():
                stats.add_skip(f"행 {idx}: 자산번호 중복 ({asset_tag})")
                continue

            # 데이터 추출 - 모든 엑셀 컬럼
            current_user_name = clean_string(row[1])  # 현 사용자
            checkout_date_str = row[2]  # 반출날짜
            return_date_str = row[3]  # 반납날짜
            status_str = clean_string(row[4])  # 상태
            previous_user_1 = clean_string(row[5])  # 이전 사용자 1
            previous_user_2 = clean_string(row[6])  # 이전 사용자 2
            first_user = clean_string(row[7])  # 최초 사용자
            location_main = clean_string(row[8])  # 위치 (판교, 대전)
            location_detail = clean_string(row[9])  # 위치 상세
            serial_number = clean_string(row[10])  # MAC 또는 시리얼넘버
            qr_code_exists = clean_string(row[11])  # QR코드 유무
            old_asset_number = clean_string(row[12])  # 기존번호
            purchase_request = clean_string(row[13])  # 구매 품의

            # 시리얼넘버 placeholder 값 처리 (unique 제약 위반 방지)
            original_serial = serial_number
            if serial_number and serial_number in ["확인필요", "미확인", "N/A", "없음", "-"]:
                serial_number = None

            # 시리얼넘버 중복 체크
            if serial_number:
                duplicate_check = await db.execute(
                    select(Asset).where(Asset.serial_number == serial_number)
                )
                if duplicate_check.scalar_one_or_none():
                    # 중복 시리얼넘버는 notes에 기록하고 NULL로 설정
                    if notes:
                        notes = f"시리얼넘버 중복: {serial_number} | {notes}"
                    else:
                        notes = f"시리얼넘버 중복: {serial_number}"
                    serial_number = None

            purchase_date_str = row[14]  # 구매연일
            tax_invoice_date_str = row[15]  # 세금계산서 발행일
            furniture_category = clean_string(row[16])  # 집기품목
            detailed_category = clean_string(row[17])  # 상세품목
            model = clean_string(row[18])  # 규격/모델명
            price = parse_price(row[19])  # 구매가
            supplier = clean_string(row[20])  # 구매처
            notes = clean_string(row[21])  # 비고
            special_notes = clean_string(row[22])  # 특이사항

            # 날짜 파싱
            purchase_date = parse_date(purchase_date_str)
            checkout_date = parse_date(checkout_date_str)
            return_date = parse_date(return_date_str)
            tax_invoice_date = parse_date(tax_invoice_date_str)

            # 등급 자동 계산
            grade = determine_grade(purchase_date)

            # 상태 매핑
            status = map_status(status_str)

            # 위치 조회/생성
            location = None
            if location_main:
                # 위치 상세에서 건물/층 정보 추출 시도
                building = None
                floor = None
                if location_detail:
                    # 예: "본사 3층" -> building="본사", floor="3층"
                    parts = location_detail.split()
                    if len(parts) >= 1:
                        building = parts[0]
                    if len(parts) >= 2:
                        floor = parts[1]

                location = await get_or_create_location(
                    db, location_main, building, floor
                )

            # 사용자 조회
            current_user = None
            if current_user_name:
                current_user = await get_user_by_name(db, current_user_name)

            # 자산 생성
            # 긴 model 정보는 description에 저장
            description = None
            if model and len(model) > 100:
                description = model
                model_short = truncate_string(model, 100)
            else:
                model_short = model

            asset = Asset(
                id=str(uuid.uuid4()),
                asset_tag=asset_tag,
                category_id=category.id if category else None,
                model=model_short,
                serial_number=truncate_string(serial_number, 100),
                status=status,
                grade=grade,
                location_id=location.id if location else None,
                assigned_to=current_user.id if current_user else None,
                # Purchase information
                purchase_price=price,
                purchase_date=purchase_date,
                purchase_request=truncate_string(purchase_request, 100),
                tax_invoice_date=tax_invoice_date,
                supplier=truncate_string(supplier, 100),
                # Category information
                furniture_category=truncate_string(furniture_category, 50),
                detailed_category=truncate_string(detailed_category, 50),
                # Usage history
                checkout_date=checkout_date,
                return_date=return_date,
                previous_user_1=truncate_string(previous_user_1, 100),
                previous_user_2=truncate_string(previous_user_2, 100),
                first_user=truncate_string(first_user, 100),
                # Additional identification
                old_asset_number=truncate_string(old_asset_number, 50),
                qr_code_exists=truncate_string(qr_code_exists, 10),
                # Notes
                description=description,
                notes=notes,
                special_notes=special_notes,
            )

            if not dry_run:
                db.add(asset)
                await db.flush()

            stats.add_success()

            # 진행 상황 출력 (100개마다)
            if stats.success_count % 100 == 0:
                print(f"   ✅ {stats.success_count}개 처리 완료...")

        except Exception as e:
            stats.add_error(idx, str(e))


async def migrate_excel(
    excel_path: str, dry_run: bool = False, clear_existing: bool = False
) -> None:
    """Excel 데이터 마이그레이션 메인 함수"""
    print("=" * 60)
    print("📦 Excel 데이터 마이그레이션")
    print("=" * 60)
    print(f"파일: {excel_path}")
    print(f"모드: {'드라이런 (실제 저장 안함)' if dry_run else '실제 마이그레이션'}")
    print(f"기존 데이터: {'삭제' if clear_existing else '유지'}")
    print("=" * 60)

    # 파일 존재 확인
    if not Path(excel_path).exists():
        print(f"❌ 파일을 찾을 수 없습니다: {excel_path}")
        return

    # Excel 파일 로드
    print("\n📂 Excel 파일 로드 중...")
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    print(f"✅ {len(wb.sheetnames)}개 시트 발견")

    stats = MigrationStats()

    async with AsyncSessionLocal() as db:
        try:
            # 기존 데이터 삭제 (옵션)
            if clear_existing and not dry_run:
                print("\n🗑️  기존 자산 데이터 삭제 중...")
                await db.execute(text("DELETE FROM assets"))
                print("✅ 삭제 완료")

            # 각 시트별 마이그레이션
            sheet_map = {
                "데스크탑(11)": ("데스크탑", "DESKTOP"),
                "노트북(12)": ("노트북", "LAPTOP"),
                "모니터(14)": ("모니터", "MONITOR"),
            }

            for sheet_name, (category_name, category_code) in sheet_map.items():
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    await migrate_sheet(
                        db, sheet, category_name, category_code, stats, dry_run
                    )

            # 커밋 또는 롤백
            if not dry_run:
                if stats.error_count == 0:
                    await db.commit()
                    print("\n✅ 모든 변경사항 저장 완료")
                else:
                    await db.rollback()
                    print("\n❌ 오류 발생으로 모든 변경사항 롤백")
            else:
                await db.rollback()
                print("\n🔍 드라이런 완료 (변경사항 저장 안함)")

            # 통계 출력
            stats.print_summary()

        except Exception as e:
            await db.rollback()
            print(f"\n❌ 마이그레이션 실패: {e}")
            raise


def main():
    parser = argparse.ArgumentParser(description="Excel 데이터 마이그레이션")
    parser.add_argument(
        "--file",
        default="자산관리 데이터(슈커톤).xlsx",
        help="Excel 파일 경로 (기본: 자산관리 데이터(슈커톤).xlsx)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="드라이런 모드 (실제 저장하지 않고 검증만)",
    )
    parser.add_argument(
        "--clear",
        action="store_true",
        help="기존 자산 데이터 삭제 후 마이그레이션",
    )

    args = parser.parse_args()

    # 프로젝트 루트로 이동 (apps/backend -> suresoft-sams)
    project_root = Path(__file__).parent.parent.parent.parent
    excel_path = project_root / args.file

    asyncio.run(migrate_excel(str(excel_path), args.dry_run, args.clear))


if __name__ == "__main__":
    main()
